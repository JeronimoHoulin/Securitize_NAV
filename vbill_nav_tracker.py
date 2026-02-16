import os
import sys
import time
import random
from datetime import datetime, timezone

import pandas as pd
from dotenv import load_dotenv
from web3 import Web3
from openpyxl import load_workbook

# ---------------------------
# Settings
# ---------------------------
load_dotenv()

RPC_URL = os.getenv("RPC_URL")
if not RPC_URL:
    api_key = os.getenv("INFURA_API_KEY") or os.getenv("api_key")
    if not api_key:
        raise Exception("Set RPC_URL or INFURA_API_KEY (or api_key) in .env")
    RPC_URL = f"https://mainnet.infura.io/v3/{api_key}"

VBILL_TOKEN_ADDRESS = "0x2255718832bc9fd3be1caf75084f4803da14ff01"
ISSUER_CONTRACT_ADDRESS = "0x22afdb66dc56be3a81285d953124bda8020dcb88"
ISSUER_CREATION_BLOCK = 22468524
NULL_ADDRESS = "0x0000000000000000000000000000000000000000"
BULK_ISSUANCE_SELECTOR = "0xb28d07c3".replace("0x", "").lower()

TEST_MODE = "--test" in sys.argv
OUTPUT_FILE = "nav_volatility_analysis_test.xlsx" if TEST_MODE else "nav_volatility_analysis.xlsx"
STRICT_MODE = os.getenv("STRICT_MODE", "1").strip() != "0"

ANALYSIS_START_DATE = datetime.strptime(
    os.getenv("ANALYSIS_START_DATE", "2025-08-21"), "%Y-%m-%d"
).date()

# Anti-429 knobs
RPC_MIN_INTERVAL_SEC = float(os.getenv("RPC_MIN_INTERVAL_SEC", "0.20"))   # ~5 req/s
RPC_MAX_RETRIES = int(os.getenv("RPC_MAX_RETRIES", "10"))
RPC_BACKOFF_BASE_SEC = float(os.getenv("RPC_BACKOFF_BASE_SEC", "0.7"))
RPC_BACKOFF_CAP_SEC = float(os.getenv("RPC_BACKOFF_CAP_SEC", "25"))

LOG_BATCH_START = int(os.getenv("LOG_BATCH_START", "3000"))
LOG_BATCH_MIN = int(os.getenv("LOG_BATCH_MIN", "250"))
LOG_BATCH_MAX = int(os.getenv("LOG_BATCH_MAX", "8000"))


def arg_value(prefix: str):
    for a in sys.argv:
        if a.startswith(prefix):
            return a.split("=", 1)[1]
    return None


if arg_value("--start-date="):
    ANALYSIS_START_DATE = datetime.strptime(arg_value("--start-date="), "%Y-%m-%d").date()
if arg_value("--rpc-min-interval="):
    RPC_MIN_INTERVAL_SEC = float(arg_value("--rpc-min-interval="))
if arg_value("--log-batch-start="):
    LOG_BATCH_START = int(arg_value("--log-batch-start="))
if arg_value("--log-batch-min="):
    LOG_BATCH_MIN = int(arg_value("--log-batch-min="))
if arg_value("--log-batch-max="):
    LOG_BATCH_MAX = int(arg_value("--log-batch-max="))


# ---------------------------
# Rate-limit + retry helpers
# ---------------------------
class RateLimiter:
    def __init__(self, min_interval_sec: float):
        self.min_interval_sec = max(0.0, float(min_interval_sec))
        self._last = 0.0

    def wait(self):
        if self.min_interval_sec <= 0:
            return
        now = time.monotonic()
        elapsed = now - self._last
        if elapsed < self.min_interval_sec:
            time.sleep(self.min_interval_sec - elapsed)
        self._last = time.monotonic()


limiter = RateLimiter(RPC_MIN_INTERVAL_SEC)


def _msg(e: Exception) -> str:
    return str(e).lower()


def is_rate_limited(e: Exception) -> bool:
    m = _msg(e)
    return any(x in m for x in [
        "429", "too many requests", "rate limit", "request rate exceeded",
        "project id request rate exceeded", "credits"
    ])


def is_too_many_results(e: Exception) -> bool:
    m = _msg(e)
    return any(x in m for x in [
        "more than 10000 results", "query returned more than", "log response size exceeded", "-32005"
    ])


def is_transient(e: Exception) -> bool:
    m = _msg(e)
    return any(x in m for x in [
        "timeout", "timed out", "temporarily unavailable", "service unavailable", "bad gateway",
        "gateway timeout", "connection aborted", "connection reset", "remote disconnected",
        "503", "502", "504", "try again"
    ])


def backoff(attempt: int, rate_limited: bool = False):
    base = RPC_BACKOFF_BASE_SEC * (2 ** max(0, attempt - 1))
    if rate_limited:
        base *= 1.4
    wait_s = min(RPC_BACKOFF_CAP_SEC, base) + random.uniform(0, 0.35)
    time.sleep(wait_s)
    return wait_s


def safe_call(label: str, fn, *args, **kwargs):
    for attempt in range(1, RPC_MAX_RETRIES + 1):
        try:
            limiter.wait()
            return fn(*args, **kwargs)
        except Exception as e:
            if not (is_rate_limited(e) or is_transient(e)) or attempt == RPC_MAX_RETRIES:
                raise RuntimeError(f"{label} failed after {attempt} attempts: {e}") from e
            waited = backoff(attempt, is_rate_limited(e))
            print(f"  Retry {attempt}/{RPC_MAX_RETRIES} for {label} ({waited:.2f}s): {e}")


def scan_transfer_logs(w3, token, topic0, topic1, topic2, start_block, end_block, label):
    """
    Adaptive scanner for Transfer logs.
    - throttles every request
    - retries with backoff
    - shrinks range on 429/result-size issues
    - no silent skipping in STRICT_MODE
    """
    all_logs, failed = [], []
    batch = max(LOG_BATCH_MIN, min(LOG_BATCH_START, LOG_BATCH_MAX))
    cur = start_block
    token = Web3.to_checksum_address(token)

    while cur <= end_block:
        to_block = min(cur + batch - 1, end_block)
        attempts = 0

        while True:
            try:
                limiter.wait()
                logs = w3.eth.get_logs({
                    "fromBlock": cur,
                    "toBlock": to_block,
                    "address": token,
                    "topics": [topic0, topic1, topic2],
                })
                all_logs.extend(logs)
                print(f"  {label} {cur}-{to_block}: +{len(logs)} (total {len(all_logs)}), batch={batch}")

                if attempts == 0 and batch < LOG_BATCH_MAX:
                    batch = min(LOG_BATCH_MAX, int(batch * 1.08) + 1)

                cur = to_block + 1
                break

            except Exception as e:
                attempts += 1
                limited = is_rate_limited(e)
                too_many = is_too_many_results(e)
                transient = is_transient(e)

                if too_many:
                    new_batch = max(LOG_BATCH_MIN, batch // 2)
                    if new_batch == batch and batch == LOG_BATCH_MIN:
                        msg = f"Unrecoverable {label} range {cur}-{to_block}: {e}"
                        if STRICT_MODE:
                            raise RuntimeError(msg) from e
                        failed.append((cur, to_block, str(e)))
                        print(f"  {msg} (strict=off, skipping)")
                        cur = to_block + 1
                        break
                    batch = new_batch
                    to_block = min(cur + batch - 1, end_block)
                    print(f"  Too many results; batch -> {batch}, retry {cur}-{to_block}")
                    continue

                if limited or transient:
                    if attempts >= 2:
                        batch = max(LOG_BATCH_MIN, batch // 2)
                        to_block = min(cur + batch - 1, end_block)

                    if attempts >= RPC_MAX_RETRIES:
                        msg = f"Failed {label} range {cur}-{to_block} after {attempts} retries: {e}"
                        if STRICT_MODE:
                            raise RuntimeError(msg) from e
                        failed.append((cur, to_block, str(e)))
                        print(f"  {msg} (strict=off, skipping)")
                        cur = to_block + 1
                        break

                    waited = backoff(attempts, limited)
                    print(f"  Retry {attempts}/{RPC_MAX_RETRIES} {label} {cur}-{to_block}, batch={batch} ({waited:.2f}s)")
                    continue

                msg = f"Non-retryable {label} range {cur}-{to_block}: {e}"
                if STRICT_MODE:
                    raise RuntimeError(msg) from e
                failed.append((cur, to_block, str(e)))
                print(f"  {msg} (strict=off, skipping)")
                cur = to_block + 1
                break

    return all_logs, failed


def decode_transfer_logs(contract_evt, raw_logs):
    out, bad = [], 0
    for lg in raw_logs:
        try:
            d = contract_evt.events.Transfer().process_log(lg)
            out.append({
                "value": int(d["args"]["value"]),
                "block": int(lg["blockNumber"]),
                "tx_hash": lg["transactionHash"].hex(),
            })
        except Exception:
            bad += 1
    return out, bad


def extract_method_id(tx_input):
    if tx_input is None:
        return ""
    if isinstance(tx_input, (bytes, bytearray)):
        return tx_input[:4].hex().lower()
    s = str(tx_input)
    return s[2:10].lower() if s.startswith("0x") else s[:8].lower()


def rolling_apy(daily_return: pd.Series, window: int) -> pd.Series:
    return (
        (1 + daily_return)
        .rolling(window=window, min_periods=window)
        .apply(lambda x: (x.prod() ** (365.0 / window)) - 1.0, raw=False)
    ) * 100.0


def since_inception_apy(df: pd.DataFrame, start_date=None):
    if df.empty:
        return None
    w = df.copy()
    w["Date_dt"] = pd.to_datetime(w["Date"], errors="coerce")
    if start_date is not None:
        w = w[w["Date_dt"].dt.date >= start_date].copy()
    w = w.sort_values("Date_dt").reset_index(drop=True)
    if len(w) < 2:
        return None

    first_d = w["Date_dt"].iloc[0].date()
    last_d = w["Date_dt"].iloc[-1].date()
    days = (last_d - first_d).days
    if days < 7:
        return None

    dr = pd.to_numeric(w["daily_return"], errors="coerce").fillna(0.0)
    growth = (1.0 + dr).prod()
    if growth <= 0:
        return None
    return (growth ** (365.0 / days) - 1.0) * 100.0


def annualised_vol(daily_std_pct: float):
    return daily_std_pct * (365 ** 0.5) if daily_std_pct is not None else None


def latest_non_null(s: pd.Series):
    s2 = pd.to_numeric(s, errors="coerce").dropna()
    return s2.iloc[-1] if len(s2) else None


def style_excel(path: str, sheet="VBILL"):
    wb = load_workbook(path)
    ws = wb[sheet]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}

    def fmt(col, nf):
        idx = hdr.get(col)
        if not idx:
            return
        for r in range(2, ws.max_row + 1):
            ws.cell(r, idx).number_format = nf

    fmt("Date", "yyyy-mm-dd")
    for c in ["Daily % change", "APY 7D", "APY 30D", "APY 90D", "APY Since Inception", "APY Since Inception (since analysis start)"]:
        fmt(c, '0.00000"%"')
    for c in ["Interest_Minted", "Supply_Minted", "Burned", "Net_Supply_Change", "Supply_Before", "Supply_After"]:
        fmt(c, "#,##0.00")
    fmt("NAV/Share", "0.000000")
    fmt("Sharpe Ratio", "0.0000")
    fmt("Block", "0")
    wb.save(path)


# ---------------------------
# Connect + contracts
# ---------------------------
w3 = Web3(Web3.HTTPProvider(RPC_URL, request_kwargs={"timeout": 120}))
if not w3.is_connected():
    raise Exception(f"RPC connect failed: {RPC_URL}")

latest_block = safe_call("latest block", lambda: w3.eth.block_number)
print(f"Connected. Latest block: {latest_block}")
print(f"429 guard: min_interval={RPC_MIN_INTERVAL_SEC}s, log_batch={LOG_BATCH_START}/{LOG_BATCH_MIN}/{LOG_BATCH_MAX}, retries={RPC_MAX_RETRIES}")

TRANSFER_ABI = {
    "anonymous": False,
    "inputs": [
        {"indexed": True, "internalType": "address", "name": "from", "type": "address"},
        {"indexed": True, "internalType": "address", "name": "to", "type": "address"},
        {"indexed": False, "internalType": "uint256", "name": "value", "type": "uint256"},
    ],
    "name": "Transfer",
    "type": "event",
}
ERC20_ABI = [
    {"constant": True, "inputs": [], "name": "decimals", "outputs": [{"name": "", "type": "uint8"}], "type": "function"},
    {"constant": True, "inputs": [], "name": "totalSupply", "outputs": [{"name": "", "type": "uint256"}], "type": "function"},
]

token_full = w3.eth.contract(address=Web3.to_checksum_address(VBILL_TOKEN_ADDRESS), abi=ERC20_ABI + [TRANSFER_ABI])
token_evt = w3.eth.contract(address=Web3.to_checksum_address(VBILL_TOKEN_ADDRESS), abi=[TRANSFER_ABI])

try:
    DECIMALS = int(safe_call("decimals", token_full.functions.decimals().call))
except Exception as e:
    print(f"Warning: decimals failed; using 6. Error: {e}")
    DECIMALS = 6

transfer_topic = Web3.to_hex(Web3.keccak(text="Transfer(address,address,uint256)"))
null_topic = "0x" + ("0" * 24) + NULL_ADDRESS[2:].lower()
issuer = Web3.to_checksum_address(ISSUER_CONTRACT_ADDRESS).lower()

# ---------------------------
# Range
# ---------------------------
from_block = int(arg_value("--from-block=") or ISSUER_CREATION_BLOCK)
if TEST_MODE:
    to_block = min(from_block + 216_000, latest_block)
    print(f"TEST mode: {from_block} -> {to_block}")
else:
    to_block = int(arg_value("--to-block=") or latest_block)
    to_block = min(to_block, latest_block)
    print(f"Scan range: {from_block} -> {to_block}")

if from_block > to_block:
    raise ValueError("from_block > to_block")

# ---------------------------
# Fetch mints
# ---------------------------
print("\nStep 1/4: mints")
mint_raw, mint_failed = scan_transfer_logs(
    w3, VBILL_TOKEN_ADDRESS, transfer_topic, null_topic, None, from_block, to_block, "mint"
)
mints, bad_mints = decode_transfer_logs(token_evt, mint_raw)
if bad_mints:
    print(f"  Warning: undecodable mint logs: {bad_mints}")
print(f"  Mint events: {len(mints)}")

# ---------------------------
# Classify mint txs
# ---------------------------
print("\nStep 2/4: classify mint txs")
tx_cache = {}
for i, m in enumerate(mints, start=1):
    txh = m["tx_hash"]
    if txh in tx_cache:
        m["is_interest"] = tx_cache[txh]
    else:
        tx = safe_call(f"tx {txh}", w3.eth.get_transaction, txh)
        tx_to = tx.get("to", None) if hasattr(tx, "get") else tx["to"]
        tx_input = tx.get("input", None) if hasattr(tx, "get") else tx["input"]
        method = extract_method_id(tx_input)
        is_interest = bool(tx_to) and str(tx_to).lower() == issuer and method == BULK_ISSUANCE_SELECTOR
        tx_cache[txh] = is_interest
        m["is_interest"] = is_interest
    if i % 100 == 0 or i == len(mints):
        print(f"  {i}/{len(mints)}")

interest_n = sum(1 for x in mints if x["is_interest"])
print(f"  Interest mints: {interest_n}, non-interest mints: {len(mints) - interest_n}")

# ---------------------------
# Fetch burns
# ---------------------------
print("\nStep 3/4: burns")
burn_raw, burn_failed = scan_transfer_logs(
    w3, VBILL_TOKEN_ADDRESS, transfer_topic, None, null_topic, from_block, to_block, "burn"
)
burns, bad_burns = decode_transfer_logs(token_evt, burn_raw)
if bad_burns:
    print(f"  Warning: undecodable burn logs: {bad_burns}")
print(f"  Burn events: {len(burns)}")

if STRICT_MODE and (mint_failed or burn_failed):
    raise RuntimeError("Failed scan ranges found in STRICT_MODE. Aborting.")

# ---------------------------
# Aggregate daily
# ---------------------------
print("\nStep 4/4: aggregate daily")
events = []
for x in mints:
    events.append({"type": "mint", "interest": x["is_interest"], "value": x["value"], "block": x["block"]})
for x in burns:
    events.append({"type": "burn", "interest": False, "value": x["value"], "block": x["block"]})
events.sort(key=lambda z: z["block"])

block_ts = {}
def get_ts(bn):
    if bn in block_ts:
        return block_ts[bn]
    b = safe_call(f"block {bn}", w3.eth.get_block, bn)
    ts = int(b["timestamp"])
    block_ts[bn] = ts
    return ts

daily = {}
for ev in events:
    d = datetime.fromtimestamp(get_ts(ev["block"]), tz=timezone.utc).date()
    if d not in daily:
        daily[d] = {"interest": 0, "minted": 0, "burned": 0, "blocks": []}
    if ev["type"] == "mint":
        if ev["interest"]:
            daily[d]["interest"] += ev["value"]
        else:
            daily[d]["minted"] += ev["value"]
    else:
        daily[d]["burned"] += ev["value"]
    daily[d]["blocks"].append(ev["block"])

if not daily:
    print("No data found.")
    sys.exit(0)

dates = pd.date_range(min(daily.keys()), max(daily.keys()), freq="D").date
rows, supply = [], 0

for d in dates:
    x = daily.get(d, {"interest": 0, "minted": 0, "burned": 0, "blocks": []})
    interest = int(x["interest"])
    minted = int(x["minted"])
    burned = int(x["burned"])

    net = interest + minted - burned
    before = supply
    after = supply + net
    dr = (interest / before) if before > 0 else 0.0

    rows.append({
        "Date": d,
        "NAV/Share": 1.0,
        "Daily % change": dr * 100.0,
        "daily_return": dr,
        "Interest_Minted": interest / (10 ** DECIMALS),
        "Supply_Minted": minted / (10 ** DECIMALS),
        "Burned": burned / (10 ** DECIMALS),
        "Net_Supply_Change": net / (10 ** DECIMALS),
        "Supply_Before": before / (10 ** DECIMALS),
        "Supply_After": after / (10 ** DECIMALS),
        "Block": max(x["blocks"]) if x["blocks"] else None,
    })
    supply = after

df = pd.DataFrame(rows)
print(f"  Days: {len(df)}")

for w in (7, 30, 90):
    df[f"APY {w}D"] = rolling_apy(pd.to_numeric(df["daily_return"], errors="coerce"), w) if len(df) >= w else pd.NA

df["APY Since Inception"] = since_inception_apy(df, None)
df["APY Since Inception (since analysis start)"] = since_inception_apy(df, ANALYSIS_START_DATE)

drs = pd.to_numeric(df["daily_return"], errors="coerce").dropna()
if len(drs) >= 7 and drs.std() and drs.std() > 0:
    df["Sharpe Ratio"] = (drs.mean() / drs.std()) * (365 ** 0.5)
else:
    df["Sharpe Ratio"] = pd.NA

# ---------------------------
# Quality check
# ---------------------------
computed_final_raw = int(round(float(df["Supply_After"].iloc[-1]) * (10 ** DECIMALS)))
onchain_final_raw = int(safe_call("totalSupply", token_full.functions.totalSupply().call, block_identifier=to_block))
diff = computed_final_raw - onchain_final_raw
quality_ok = (diff == 0) and not mint_failed and not burn_failed

print("\nQuality check")
print(f"  Computed final supply: {computed_final_raw / (10 ** DECIMALS):,.6f}")
print(f"  On-chain totalSupply : {onchain_final_raw / (10 ** DECIMALS):,.6f}")
print(f"  Difference           : {diff / (10 ** DECIMALS):,.6f}")
print(f"  Failed mint ranges   : {len(mint_failed)}")
print(f"  Failed burn ranges   : {len(burn_failed)}")

if STRICT_MODE and not quality_ok:
    raise RuntimeError("Quality check failed in STRICT_MODE. Aborting export.")

# ---------------------------
# Export (replace only VBILL sheet)
# ---------------------------
excel_df = df.copy()
excel_df["Date"] = pd.to_datetime(excel_df["Date"], errors="coerce")
for c in [
    "NAV/Share", "Daily % change", "daily_return", "Interest_Minted", "Supply_Minted", "Burned",
    "Net_Supply_Change", "Supply_Before", "Supply_After",
    "APY 7D", "APY 30D", "APY 90D",
    "APY Since Inception", "APY Since Inception (since analysis start)",
    "Sharpe Ratio", "Block"
]:
    if c in excel_df.columns:
        excel_df[c] = pd.to_numeric(excel_df[c], errors="coerce")

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=False)
    old_sheets = wb.sheetnames
    wb.close()
    print(f"\nWorkbook exists. Sheets: {', '.join(old_sheets)}")
else:
    old_sheets = []

print(f"Exporting to {OUTPUT_FILE} ...")
if os.path.exists(OUTPUT_FILE):
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        excel_df.to_excel(w, sheet_name="VBILL", index=False)
else:
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", mode="w") as w:
        excel_df.to_excel(w, sheet_name="VBILL", index=False)

style_excel(OUTPUT_FILE, "VBILL")
print(f"Updated sheet 'VBILL' ({len(excel_df)} rows)")
if old_sheets:
    untouched = [s for s in old_sheets if s != "VBILL"]
    if untouched:
        print(f"Untouched sheets: {', '.join(untouched)}")

# ---------------------------
# Summary
# ---------------------------
daily_changes = pd.to_numeric(df["Daily % change"], errors="coerce").dropna()
daily_std = float(daily_changes.std()) if len(daily_changes) else None
ann_vol = annualised_vol(daily_std) if daily_std is not None else None

print("\n" + "=" * 98)
print("VBILL NAV TRACKING SUMMARY")
print("=" * 98)
print(f"Days tracked (calendar): {len(df)}")
print(f"Mint events: {len(mints)}")
print(f"Burn events: {len(burns)}")
print(f"Total interest minted: {float(df['Interest_Minted'].sum()):,.2f}")
print(f"Total supply minted  : {float(df['Supply_Minted'].sum()):,.2f}")
print(f"Total burned         : {float(df['Burned'].sum()):,.2f}")
print(f"Final supply (computed): {float(df['Supply_After'].iloc[-1]):,.2f}")
print(f"Final supply (on-chain): {onchain_final_raw / (10 ** DECIMALS):,.2f}")
print(f"Max loss: {float(daily_changes.min()):.5f}%" if len(daily_changes) else "Max loss: n/a")
print(f"Daily std dev: {daily_std:.8f}%" if daily_std is not None else "Daily std dev: n/a")
print(f"Annualised volatility: {ann_vol:.4f}%" if ann_vol is not None else "Annualised volatility: n/a")

apy7 = latest_non_null(df["APY 7D"])
apy30 = latest_non_null(df["APY 30D"])
apy90 = latest_non_null(df["APY 90D"])
apy_full = latest_non_null(df["APY Since Inception"])
apy_from_date = latest_non_null(df["APY Since Inception (since analysis start)"])
sharpe = latest_non_null(df["Sharpe Ratio"])

print(f"Latest APY 7D: {apy7:.3f}%" if apy7 is not None else "Latest APY 7D: n/a")
print(f"Latest APY 30D: {apy30:.3f}%" if apy30 is not None else "Latest APY 30D: n/a")
print(f"Latest APY 90D: {apy90:.3f}%" if apy90 is not None else "Latest APY 90D: n/a")
print(f"APY since inception (full): {apy_full:.3f}%" if apy_full is not None else "APY since inception (full): n/a")
print(
    f"APY since inception (since {ANALYSIS_START_DATE}): {apy_from_date:.3f}%"
    if apy_from_date is not None else f"APY since inception (since {ANALYSIS_START_DATE}): n/a"
)
print(f"Sharpe ratio: {sharpe:.4f}" if sharpe is not None else "Sharpe ratio: n/a")
print(f"Data quality status: {'PASS' if quality_ok else 'FAIL'}")
if not quality_ok:
    print("Warning: quality checks failed. Review failed ranges and/or RPC reliability.")
print()
